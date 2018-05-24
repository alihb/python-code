# python-code
from openpyxl import load_workbook

wb = load_workbook('expence.xlsx')
# wb.create_sheet('exlist')
ws = wb['exlist']

ws['A1'].value = 'Code'
ws['B1'].value = 'Name'
ws['C1'].value = 'Expence'
ws['D1'].value = 'Date'

ws.append([ raw_input('Code:', ),
            raw_input('Name:', ),
            raw_input('Expe:', ),
            raw_input('Date:', )])

wb.save('expence.xlsx')
